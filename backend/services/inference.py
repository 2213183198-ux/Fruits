from __future__ import annotations

import csv
import re
import time
from collections import Counter
from pathlib import Path
from typing import Any, Dict, Iterable, Optional

import cv2
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from ultralytics import YOLO

from backend.config import settings
from backend.schemas import QualityRuleSettings
from backend.services.quality_rules import build_default_quality_rule_settings


class InferenceService:
    def __init__(self, model_path: Path) -> None:
        self.model_path = model_path
        self._model: Optional[YOLO] = None
        self._loaded_model_path: Optional[Path] = None
        self._class_names_override: list[str] = []
        self._quality_rule_settings: QualityRuleSettings = build_default_quality_rule_settings()

    @property
    def model_loaded(self) -> bool:
        return self._model is not None

    def load_model(self) -> YOLO:
        current_path = self.model_path.resolve()
        if self._model is None or self._loaded_model_path != current_path:
            if not self.model_path.exists():
                raise FileNotFoundError(f"Model file not found: {self.model_path}")
            if self.model_path.suffix.lower() == ".onnx":
                self._model = YOLO(str(self.model_path), task="detect")
            else:
                self._model = YOLO(str(self.model_path))
            self._loaded_model_path = current_path
        return self._model

    def set_model_path(self, model_path: Path) -> None:
        resolved = model_path.resolve()
        if resolved == self.model_path.resolve():
            return
        self.model_path = resolved
        self._model = None
        self._loaded_model_path = None

    def set_class_names_override(self, class_names: list[str] | None) -> None:
        self._class_names_override = [str(item) for item in (class_names or [])]

    def set_quality_rule_settings(self, quality_rule_settings: QualityRuleSettings) -> None:
        self._quality_rule_settings = QualityRuleSettings.model_validate(quality_rule_settings.model_dump())

    def get_model_name(self) -> str:
        return self.model_path.name

    def get_model_type(self) -> str:
        suffix = self.model_path.suffix.lower()
        if suffix == ".onnx":
            return "onnx"
        if suffix == ".pt":
            return "pytorch"
        return suffix.lstrip(".") or "unknown"

    def predict_bytes(
        self,
        file_bytes: bytes,
        filename: str,
        imgsz: int,
        conf: float,
        save_artifact: bool,
    ) -> Dict[str, Any]:
        image = self._decode_image(file_bytes)
        model = self.load_model()
        results = model.predict(source=image, imgsz=imgsz, conf=conf, verbose=False)
        result = results[0]

        detections: list[dict[str, Any]] = []
        class_counts: Counter[str] = Counter()
        fresh_count = 0
        rotten_count = 0

        if result.boxes is not None:
            for box in result.boxes:
                class_id = int(box.cls.item())
                label = self._resolve_label(result.names, class_id)
                confidence = round(float(box.conf.item()), 4)
                x1, y1, x2, y2 = [round(float(value), 2) for value in box.xyxy[0].tolist()]

                detections.append(
                    {
                        "class_id": class_id,
                        "label": label,
                        "confidence": confidence,
                        "bbox": {"x1": x1, "y1": y1, "x2": x2, "y2": y2},
                    }
                )
                class_counts[label] += 1
                if self._matches_keywords(label, self._quality_rule_settings.rotten_keywords):
                    rotten_count += 1
                elif self._matches_keywords(label, self._quality_rule_settings.fresh_keywords):
                    fresh_count += 1

        artifact_url = None
        if save_artifact:
            artifact_url = self._save_artifact(result.plot(), filename)

        total_count = len(detections)
        report_payload = self._build_report(
            class_counts=class_counts,
            total_count=total_count,
            fresh_count=fresh_count,
            rotten_count=rotten_count,
        )

        return {
            "filename": filename,
            "model_name": self.get_model_name(),
            "model_path": str(self.model_path),
            "model_type": self.get_model_type(),
            "quality_rule_applied": report_payload["quality_rule_applied"],
            "image_size": imgsz,
            "confidence_threshold": conf,
            "detections": detections,
            "summary": {
                "total_detections": len(detections),
                "class_counts": dict(sorted(class_counts.items())),
            },
            "report": report_payload["report"],
            "artifact_url": artifact_url,
        }

    def _decode_image(self, file_bytes: bytes) -> np.ndarray:
        buffer = np.frombuffer(file_bytes, dtype=np.uint8)
        image = cv2.imdecode(buffer, cv2.IMREAD_COLOR)
        if image is None:
            raise ValueError("Uploaded file is not a valid image.")
        return image

    def _save_artifact(self, annotated_image: np.ndarray, filename: str) -> str:
        settings.artifacts_dir.mkdir(parents=True, exist_ok=True)
        safe_stem = re.sub(r"[^a-zA-Z0-9_-]+", "_", Path(filename).stem).strip("_") or "image"
        artifact_name = f"{safe_stem}_{int(time.time() * 1000)}.jpg"
        artifact_path = settings.artifacts_dir / artifact_name
        cv2.imwrite(str(artifact_path), annotated_image)
        return f"/artifacts/{artifact_name}"

    def export_batch_csv(self, results: Iterable[Dict[str, Any]]) -> str:
        settings.artifacts_dir.mkdir(parents=True, exist_ok=True)
        csv_name = f"batch_report_{int(time.time() * 1000)}.csv"
        csv_path = settings.artifacts_dir / csv_name

        with csv_path.open("w", newline="", encoding="utf-8-sig") as csv_file:
            writer = csv.DictWriter(
                csv_file,
                fieldnames=[
                    "filename",
                    "total_detections",
                    "fresh_count",
                    "rotten_count",
                    "rotten_rate",
                    "status",
                    "recommendation",
                    "class_counts",
                    "artifact_url",
                ],
            )
            writer.writeheader()
            for item in results:
                writer.writerow(
                    {
                        "filename": item["filename"],
                        "total_detections": item["summary"]["total_detections"],
                        "fresh_count": item["report"]["fresh_count"],
                        "rotten_count": item["report"]["rotten_count"],
                        "rotten_rate": item["report"]["rotten_rate"],
                        "status": item["report"]["status"],
                        "recommendation": item["report"]["recommendation"],
                        "class_counts": "; ".join(
                            f"{label}:{count}" for label, count in item["summary"]["class_counts"].items()
                        ),
                        "artifact_url": item.get("artifact_url") or "",
                    }
                )

        return f"/artifacts/{csv_name}"

    def export_batch_excel(
        self,
        results: Iterable[Dict[str, Any]],
        failures: Iterable[Dict[str, Any]],
        total_files: int,
    ) -> str:
        result_list = list(results)
        failure_list = list(failures)

        settings.artifacts_dir.mkdir(parents=True, exist_ok=True)
        excel_name = f"batch_report_{int(time.time() * 1000)}.xlsx"
        excel_path = settings.artifacts_dir / excel_name

        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Summary"

        status_counts: Counter[str] = Counter(item["report"]["status"] for item in result_list)
        total_detections = sum(item["summary"]["total_detections"] for item in result_list)
        total_rotten = sum(item["report"]["rotten_count"] for item in result_list)
        total_fresh = sum(item["report"]["fresh_count"] for item in result_list)

        summary_rows = [
            ("Generated At", time.strftime("%Y-%m-%d %H:%M:%S")),
            ("Total Files", total_files),
            ("Successful Files", len(result_list)),
            ("Failed Files", len(failure_list)),
            ("Total Detections", total_detections),
            ("Fresh Count", total_fresh),
            ("Rotten Count", total_rotten),
            ("Pass Count", status_counts.get("pass", 0)),
            ("Warning Count", status_counts.get("warning", 0)),
            ("Critical Count", status_counts.get("critical", 0)),
            ("Detected Only Count", status_counts.get("detected", 0)),
            ("No Detection Count", status_counts.get("no_detection", 0)),
        ]

        for row_index, (label, value) in enumerate(summary_rows, start=1):
            summary_sheet.cell(row=row_index, column=1, value=label)
            summary_sheet.cell(row=row_index, column=2, value=value)

        self._style_header(summary_sheet, [1, 2], 1)
        summary_sheet.column_dimensions["A"].width = 24
        summary_sheet.column_dimensions["B"].width = 28

        detail_sheet = workbook.create_sheet("Inspection Results")
        detail_headers = [
            "Filename",
            "Total Detections",
            "Fresh Count",
            "Rotten Count",
            "Rotten Rate",
            "Status",
            "Recommendation",
            "Class Counts",
            "Artifact URL",
        ]
        detail_sheet.append(detail_headers)
        self._style_header(detail_sheet, range(1, len(detail_headers) + 1), 1)

        for item in result_list:
            detail_sheet.append(
                [
                    item["filename"],
                    item["summary"]["total_detections"],
                    item["report"]["fresh_count"],
                    item["report"]["rotten_count"],
                    item["report"]["rotten_rate"],
                    item["report"]["status"],
                    item["report"]["recommendation"],
                    "; ".join(f"{label}:{count}" for label, count in item["summary"]["class_counts"].items()),
                    item.get("artifact_url") or "",
                ]
            )

        detection_sheet = workbook.create_sheet("Detections")
        detection_headers = [
            "Filename",
            "Label",
            "Confidence",
            "x1",
            "y1",
            "x2",
            "y2",
        ]
        detection_sheet.append(detection_headers)
        self._style_header(detection_sheet, range(1, len(detection_headers) + 1), 1)

        for item in result_list:
            for detection in item["detections"]:
                detection_sheet.append(
                    [
                        item["filename"],
                        detection["label"],
                        detection["confidence"],
                        detection["bbox"]["x1"],
                        detection["bbox"]["y1"],
                        detection["bbox"]["x2"],
                        detection["bbox"]["y2"],
                    ]
                )

        if failure_list:
            failure_sheet = workbook.create_sheet("Failures")
            failure_headers = ["Filename", "Error"]
            failure_sheet.append(failure_headers)
            self._style_header(failure_sheet, range(1, len(failure_headers) + 1), 1)
            for item in failure_list:
                failure_sheet.append([item["filename"], item["error"]])

        for sheet in workbook.worksheets:
            self._autosize_sheet(sheet)

        workbook.save(excel_path)
        return f"/artifacts/{excel_name}"

    def _resolve_label(self, names: Any, class_id: int) -> str:
        if self._class_names_override and 0 <= class_id < len(self._class_names_override):
            return self._class_names_override[class_id]
        if isinstance(names, dict):
            return str(names.get(class_id, class_id))
        if isinstance(names, list) and 0 <= class_id < len(names):
            return str(names[class_id])
        return str(class_id)

    def _style_header(self, sheet, columns, row_index: int) -> None:
        fill = PatternFill(fill_type="solid", start_color="F6EAD7", end_color="F6EAD7")
        font = Font(bold=True)
        for column in columns:
            cell = sheet.cell(row=row_index, column=column)
            cell.fill = fill
            cell.font = font

    def _autosize_sheet(self, sheet) -> None:
        for column_cells in sheet.columns:
            values = [len(str(cell.value)) for cell in column_cells if cell.value is not None]
            column_letter = column_cells[0].column_letter
            sheet.column_dimensions[column_letter].width = min(max(values, default=12) + 2, 48)

    def _build_report(
        self,
        class_counts: Counter[str],
        total_count: int,
        fresh_count: int,
        rotten_count: int,
    ) -> Dict[str, Any]:
        quality_rules = self._quality_rule_settings
        quality_rule_applied = self._supports_quality_rule(class_counts)
        rotten_rate = round((rotten_count / total_count), 4) if total_count else 0.0

        if total_count == 0:
            return {
                "quality_rule_applied": quality_rule_applied,
                "report": {
                    "fresh_count": fresh_count,
                    "rotten_count": rotten_count,
                    "rotten_rate": rotten_rate,
                    "status": "no_detection",
                    "recommendation": quality_rules.messages.no_detection,
                },
            }

        if not quality_rule_applied:
            return {
                "quality_rule_applied": False,
                "report": {
                    "fresh_count": 0,
                    "rotten_count": 0,
                    "rotten_rate": 0.0,
                    "status": "detected",
                    "recommendation": quality_rules.messages.detected_only,
                },
            }

        if rotten_rate <= quality_rules.pass_max_rotten_rate:
            status = "pass"
            recommendation = quality_rules.messages.pass_message
        elif rotten_rate <= quality_rules.warning_max_rotten_rate:
            status = "warning"
            recommendation = quality_rules.messages.warning_message
        else:
            status = "critical"
            recommendation = quality_rules.messages.critical_message

        return {
            "quality_rule_applied": True,
            "report": {
                "fresh_count": fresh_count,
                "rotten_count": rotten_count,
                "rotten_rate": rotten_rate,
                "status": status,
                "recommendation": recommendation,
            },
        }

    def _supports_quality_rule(self, class_counts: Counter[str]) -> bool:
        if not self._quality_rule_settings.enabled:
            return False
        if not class_counts:
            return True

        fresh_keywords = self._quality_rule_settings.fresh_keywords
        rotten_keywords = self._quality_rule_settings.rotten_keywords
        has_fresh = any(self._matches_keywords(label, fresh_keywords) for label in class_counts.keys())
        has_rotten = any(self._matches_keywords(label, rotten_keywords) for label in class_counts.keys())
        return has_fresh or has_rotten

    def _matches_keywords(self, label: str, keywords: list[str]) -> bool:
        lowered_label = str(label).casefold()
        return any(str(keyword).strip().casefold() in lowered_label for keyword in keywords if str(keyword).strip())


inference_service = InferenceService(settings.model_path)
